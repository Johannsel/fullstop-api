import requests
import time
import json
import openpyxl
import atexit
from openpyxl import Workbook

def dev_request(to_process, to_compare, fstop_threshold, url):
    global_headers = {'Content-Type':'application/json'}
    data = {'to_process': to_process, 'to_compare': to_compare, 'fstop_threshold': fstop_threshold}
    payload = dict(data = data)
    payload = json.dumps(payload)
    response = requests.post(str(url), data=payload, headers=global_headers)
    data = json.loads(response.content)
    o = data['data']
    return o

def prod_request(to_process, model ,url):
    global_headers = {'Content-Type':'application/json'}
    data = {'to_process': to_process, 'model': model}
    payload = dict(data = data)
    payload = json.dumps(payload)
    response = requests.post(str(url), data=payload, headers=global_headers, timeout=300)
    data = json.loads(response.content)
    o = data['data']
    return o

def dev_call(input_wb, output_wb, model, url, max_row):
    file = openpyxl.load_workbook(input_wb)
    sheet = file['Tabelle1']
    o = Workbook()
    o_sheet = o.active
    o_sheet["A1"] = "IDEAL:"
    o_sheet["B1"] = "PUNCTALL:"
    o_sheet["C1"] = "PUNCTALL CER:"
    o_sheet["D1"] = "FSTOP:"
    o_sheet["E1"] = "FSTOP CER:"
    o_sheet["F1"] = "ORIGINAL:"
    o_sheet["G1"] = "ORIGINAL CER:"

    for c, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, max_row=max_row,  values_only=True), start=2):
        gold_standard = row[0]
        to_process = row[1]
        fstop_threshold = 20
        print("Working for:" + str(gold_standard))
        while True:
            try:
                time.sleep(10)
                call = dev_request(to_process, gold_standard, fstop_threshold ,url)
            except:
                print("Error! Retrying in 10s.")
                continue
            break
        print("Finished")
        o_sheet[f"A{c}"] = call['ideal']
        o_sheet[f"B{c}"] = call['punctall'][0]
        o_sheet[f"C{c}"] = call['punctall'][1]
        o_sheet[f"D{c}"] = call['fstop'][0]
        o_sheet[f"E{c}"] = call['fstop'][1]
        o_sheet[f"F{c}"] = call['original'][0]
        o_sheet[f"G{c}"] = call['original'][1]

    o.save(str(output_wb))

def prod_call(input_wb, output_wb, model, url, max_row):
    file = openpyxl.load_workbook(str(input_wb))
    sheet = file['Tabelle1']
    o = Workbook()
    o_sheet = o.active
    o_sheet["A1"] = "ORIGINAL:"
    o_sheet["B1"] = "PROCESSED:"
    o_sheet["C1"] = "MODEL:"

    for c, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, max_row=max_row,  values_only=True), start=2):
        gold_standard = row[0]
        to_process = row[1]
        print("Working for:" + str(gold_standard))
        while True:
            try:
                time.sleep(10)
                call = prod_request(to_process, model, url)
            except:
                print("Error! Retrying in 10s.")
                continue
            break
        print("Finished with:" + str(call['processed'][1]))
        o_sheet[f"A{c}"] = call['processed'][0]
        o_sheet[f"B{c}"] = call['processed'][1]
        o_sheet[f"C{c}"] = call['processed'][2]

    o.save(str(output_wb))

    #@atexit.register
    #def exit_handler():
    #    print("Storing Excel on manual exit!")
    #    o.save(str(output_wb))

